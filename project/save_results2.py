import pandas as pd
from pyomo.environ import *
from time import *
from openpyxl import load_workbook
import openpyxl



def save_results2(m: ConcreteModel(), h: int, case_nr: int, prices: dict, resources: dict, number_resources: int) -> None:
    results_dict = {}
    for var in m.component_data_objects(Var, active=True):
        results_dict[var.name] = var.value

    '''# Populate DataFrame with variable values
    results = pd.DataFrame()
    v_vector = [j for j in m.component_objects(Var, active=True)]
    var_vector = [j for j in m.component_data_objects(Var, active=True)]
    print(len(v_vector))
    print(len(var_vector))
    for i in range(0, len([j for j in m.component_data_objects(Var, active=True)])):
        v = m.component_objects(Var, active=True)[i]
        var = m.component_data_objects(Var, active=True)[i]
        for index in v:
            print(v.name, var.name)
            #results.at[index, v.name] = value(v[index])'''

    results = pd.DataFrame()
    for v in m.component_objects(Var, active=True):
        for index in v:
            try:
                results.at[index, v.name] = value(v[index], exception=False)
            except:
                1

    flag_excel = 1
    iter = 0
    while flag_excel > 0:
        try:
            results.to_excel(f'Data/results_v{iter}.xlsx', float_format='%.1f')
            flag_excel = 0
        except:
            iter = iter + 1

        excel_colouring(f'Data/results_v{iter}.xlsx')

def excel_colouring(excel_file):
    # Load the Excel file
    wb = load_workbook(excel_file)
    ws = wb.active

    import string

    list_letters = list(string.ascii_uppercase)
    list_selected_letters = list(string.ascii_uppercase)
    for i in range(0, 3):
        for j in list_letters:
            list_selected_letters.append(list_letters[i] + j)
            #print(list_selected_letters)

    # Apply colors to the specified columns
    columns_to_color = ['G', 'H', 'I', 'W', ]
    list_selected_letters_zero = [i + '1' for i in list_selected_letters]
    for i in range(0, len(list_selected_letters_zero)):
        column = list_selected_letters[i]
        column_zero = ws[list_selected_letters_zero[i]].value
        if column_zero is not None:
            ws[list_selected_letters_zero[i]].alignment = ws[list_selected_letters_zero[i]].alignment.copy(wrap_text=True)
            for cell in ws[column]:
                if any(string in column_zero for string in ['PV', 'EL']):
                    cell.fill = openpyxl.styles.PatternFill(start_color='a4ffa4', end_color='a4ffa4', fill_type='solid')

    # Save the changes to the Excel file
    wb.save(excel_file)
